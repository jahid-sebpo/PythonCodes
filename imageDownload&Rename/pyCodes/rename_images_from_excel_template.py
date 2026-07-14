import os
import shutil
import openpyxl
import requests

def download_images_from_links(excel_filename="renamedImagesTemplate.xlsx", excel_folder="file", target_folder="img", sheet_name="imglink"):
    """
    Reads an Excel template from a specific sheet ('imglink'), looks up target filenames 
    in Column A and image URLs in Column B, then downloads and saves them.
    
    Args:
        excel_filename (str): The name of the Excel file.
        excel_folder (str): The folder containing the Excel file.
        target_folder (str): The folder where downloaded images should be saved.
        sheet_name (str): The name of the worksheet containing the links.
    """
    excel_path = os.path.join(excel_folder, excel_filename)
    
    # Validation checks
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at '{excel_path}'")
        return
        
    # Ensure the output directory exists
    os.makedirs(target_folder, exist_ok=True)

    try:
        workbook = openpyxl.load_workbook(excel_path)
        
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in '{excel_filename}'.")
            print(f"Available sheets: {workbook.sheetnames}")
            return
            
        sheet = workbook[sheet_name]
        print(f"Successfully loaded sheet '{sheet_name}' from '{excel_path}'")
        
        download_count = 0
        failed_count = 0
        skipped_count = 0

        # Define headers to prevent some websites from blocking our request (403 Forbidden)
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }

        # Iterate starting from Row 2 to skip headers
        for row_idx in range(2, sheet.max_row + 1):
            file_name_raw = sheet.cell(row=row_idx, column=1).value
            url_raw = sheet.cell(row=row_idx, column=2).value
            
            # Skip empty rows
            if not file_name_raw or not url_raw:
                continue
                
            file_name = str(file_name_raw).strip()
            url = str(url_raw).strip()
            
            # Fallback extension if the provided file name has no extension
            _, ext = os.path.splitext(file_name)
            if not ext:
                # Try to extract the extension from the URL, defaulting to .jpg
                parsed_ext = os.path.splitext(url.split("?")[0])[1]
                file_name = f"{file_name}{parsed_ext if parsed_ext else '.jpg'}"

            destination_path = os.path.join(target_folder, file_name)

            print(f"Downloading: {url} -> {destination_path}")
            
            try:
                # Use stream=True to handle large files efficiently
                with requests.get(url, headers=headers, stream=True, timeout=15) as response:
                    response.raise_for_status()
                    
                    with open(destination_path, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                                
                print(f"Successfully downloaded: '{file_name}'")
                download_count += 1
                
            except requests.exceptions.RequestException as req_err:
                print(f"Failed to download image from {url}. Error: {req_err}")
                failed_count += 1
            except Exception as file_err:
                print(f"Failed to save image '{file_name}'. Error: {file_err}")
                failed_count += 1

        print("\n--- Download Summary ---")
        print(f"Successfully downloaded: {download_count} files")
        print(f"Failed downloads: {failed_count} files")

    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")


def rename_and_copy_images(excel_filename="renamedImagesTemplate.xlsx", excel_folder="file", image_folder="img", target_folder="renamedImg"):
    """
    Reads an Excel template to find image names in Column A, 
    renames the matching files in the 'img' directory to the names in Column B,
    and additionally saves a copy of the renamed image to the 'renamedImg' directory.
    """
    excel_path = os.path.join(excel_folder, excel_filename)
    
    # Validation checks for paths
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at '{excel_path}'")
        return
        
    if not os.path.exists(image_folder):
        print(f"Error: Image folder '{image_folder}' does not exist.")
        return

    # Ensure the target folder for the additional copies exists
    os.makedirs(target_folder, exist_ok=True)

    try:
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        print(f"Successfully loaded sheet '{sheet.title}' from '{excel_path}'")
        
        rename_count = 0
        missing_count = 0
        skipped_count = 0

        # Iterate starting from Row 2 to skip headers
        for row_idx in range(2, sheet.max_row + 1):
            old_name_raw = sheet.cell(row=row_idx, column=1).value
            new_name_raw = sheet.cell(row=row_idx, column=2).value
            
            # Skip empty rows
            if not old_name_raw or not new_name_raw:
                continue
                
            old_name = str(old_name_raw).strip()
            new_name = str(new_name_raw).strip()
            
            # Construct the source path
            old_image_path = os.path.join(image_folder, old_name)
            
            # Check if the target source image actually exists
            if os.path.exists(old_image_path):
                # Extract the extension from the original file just in case Column B doesn't include it
                _, old_ext = os.path.splitext(old_name)
                _, new_ext = os.path.splitext(new_name)
                
                # If Column B does not have an extension, preserve the original one
                final_new_name = new_name
                if not new_ext:
                    final_new_name = f"{new_name}{old_ext}"
                    
                new_image_path = os.path.join(image_folder, final_new_name)
                copied_image_path = os.path.join(target_folder, final_new_name)
                
                # Perform the rename and copy operations
                try:
                    # 1. Rename the original file in 'img'
                    os.rename(old_image_path, new_image_path)
                    
                    # 2. Copy the renamed file to the 'renamedImg' folder
                    shutil.copy2(new_image_path, copied_image_path)
                    
                    print(f"Processed: '{old_name}' -> Renamed in '{image_folder}/' & Copied to '{target_folder}/{final_new_name}'")
                    rename_count += 1
                except Exception as file_err:
                    print(f"Failed to process '{old_name}': {file_err}")
                    skipped_count += 1
            else:
                print(f"Warning: File '{old_name}' not found in '{image_folder}/'")
                missing_count += 1

        print("\n--- Processing Summary ---")
        print(f"Successfully processed: {rename_count} files (renamed and copied)")
        print(f"Missing source files: {missing_count} files")
        if skipped_count > 0:
            print(f"Errors encountered: {skipped_count} files")

    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")

# --- Main Execution ---
# if __name__ == "__main__":
    # WORKFLOW 1: Download images from URLs listed in sheet 'imglink' (Column B)
    # and save them to the 'img' directory using names from Column A.
    #download_images_from_links()
    
    # WORKFLOW 2 (Optional/Original): Rename and copy local files
    # Uncomment the line below if you also want to execute the renaming process.
    # rename_and_copy_images()